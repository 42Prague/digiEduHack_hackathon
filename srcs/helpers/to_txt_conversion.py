import pdfplumber
import mammoth
import markdown
import requests
from bs4 import BeautifulSoup
import openpyxl
from .audio_transcription import transcribe_audio



def convert_to_text(input_path):
    """
    Converts PDF, DOCX, MD, or MP3 into a single string.
    Returns the extracted plain text.
    """
    ext = input_path.lower().split(".")[-1]
    text = ""

    # PDF → Text
    if ext == "pdf":
        with pdfplumber.open(input_path) as pdf:
            for page in pdf.pages:
                text += page.extract_text() or ""
                text += "\n"

    # DOCX → Text (Mammoth)
    elif ext == "docx":
        with open(input_path, "rb") as f:
            result = mammoth.extract_raw_text(f)
            text = result.value

    # MD → Text
    elif ext == "md":
        with open(input_path, "r", encoding="utf-8") as f:
            md = f.read()
        html = markdown.markdown(md)
        text = BeautifulSoup(html, "html.parser").get_text()

    # MP3 → Transcribed Text
    elif ext == "mp3":
        # Upload local MP3 to AssemblyAI first
        base_url = "https://api.assemblyai.com/v2/upload"
        headers = {"authorization": "6e39c034c7f3444a862e60a3245da366"}
        with open(input_path, "rb") as f:
            upload_response = requests.post(base_url, headers=headers, data=f)
            upload_response.raise_for_status()
            audio_url = upload_response.json()["upload_url"]

        # Call your transcription function with the uploaded URL
        text = transcribe_audio(audio_url)
        print(text)
    elif ext == "xlsx":
        wb = openpyxl.load_workbook(input_path, data_only=True)
        out = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            out.append(f"\n=== Sheet: {sheet} ===\n")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join("" if v is None else str(v) for v in row)
                out.append(row_text)
        text = "\n".join(out)

    else:
        raise ValueError("Unsupported file type. Use PDF, DOCX, MD, or MP3.")

    return text
